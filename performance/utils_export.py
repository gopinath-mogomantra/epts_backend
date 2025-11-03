# ===========================================================
# performance/views_reports.py (PRODUCTION-READY VERSION)
# ===========================================================
from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.db import transaction
from django.utils import timezone
import logging

from .models import PerformanceEvaluation
from employee.models import Employee, Department
from .serializers import PerformanceEvaluationSerializer


logger = logging.getLogger("performance")


# ===========================================================
# Pagination for Reports
# ===========================================================
class PerformanceReportPagination(PageNumberPagination):
    """Custom pagination for performance reports."""
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 500


# ===========================================================
# Weekly / Department / Manager Report (Optimized)
# ===========================================================
class PerformanceReportView(generics.ListAPIView):
    """
    List performance evaluations with filtering and role-based access.
    
    Query Parameters:
        - filter: 'weekly', 'department', or 'manager'
        - value: corresponding value for the filter
        - page: page number
        - page_size: results per page
    
    Role-based filtering:
        - Admin: See all evaluations
        - Manager: See only direct reports
        - Employee: See only own evaluations
    """
    serializer_class = PerformanceEvaluationSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = PerformanceReportPagination

    def get_queryset(self):
        """Get queryset with proper optimizations and filters."""
        user = self.request.user
        
        # Optimized query with all required FK relationships
        qs = PerformanceEvaluation.objects.select_related(
            "employee__user",
            "employee__department",
            "employee__manager__user",
            "evaluator",
            "department"
        )

        # Role-based filtering
        if user.role == "Manager":
            qs = qs.filter(employee__manager__user=user)
        elif user.role == "Employee":
            qs = qs.filter(employee__user=user)
        # Admin sees all (no filter)

        # Apply additional filters
        filter_type = self.request.query_params.get("filter", "").lower()
        value = self.request.query_params.get("value")

        if filter_type == "weekly" and value:
            try:
                week = int(value)
                if 1 <= week <= 53:
                    qs = qs.filter(week_number=week)
            except (ValueError, TypeError):
                logger.warning(f"Invalid week number: {value}")
                
        elif filter_type == "department" and value:
            qs = qs.filter(department__code__iexact=value)
            
        elif filter_type == "manager" and value:
            qs = qs.filter(employee__manager__user__emp_id__iexact=value)
        
        # Additional filters
        year = self.request.query_params.get("year")
        if year:
            try:
                qs = qs.filter(year=int(year))
            except (ValueError, TypeError):
                pass

        return qs.order_by("-year", "-week_number", "-average_score")


# ===========================================================
# Excel Export (Memory-Optimized)
# ===========================================================
class PerformanceExcelExportView(generics.GenericAPIView):
    """
    Export performance evaluations to Excel with filtering.
    
    Query Parameters:
        - filter: 'department', 'manager', or 'week'
        - value: corresponding value for the filter
        - year: filter by year
    
    Limits: Max 10,000 records per export to prevent memory issues.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    MAX_EXPORT_SIZE = 10000

    @transaction.atomic
    def get(self, request):
        """Generate Excel export with proper optimization."""
        user = request.user
        
        # Base queryset with all required FKs
        qs = PerformanceEvaluation.objects.select_related(
            "employee__user",
            "employee__department",
            "employee__manager__user",
            "department"
        )
        
        # Role-based filtering
        if user.role == "Manager":
            qs = qs.filter(employee__manager__user=user)
        elif user.role == "Employee":
            qs = qs.filter(employee__user=user)
        
        # Apply filters
        filter_type = request.query_params.get("filter")
        value = request.query_params.get("value")

        if filter_type == "department" and value:
            qs = qs.filter(department__code__iexact=value)
        elif filter_type == "manager" and value:
            qs = qs.filter(employee__manager__user__emp_id__iexact=value)
        elif filter_type == "week" and value:
            try:
                qs = qs.filter(week_number=int(value))
            except (ValueError, TypeError):
                pass
        
        # Year filter
        year = request.query_params.get("year")
        if year:
            try:
                qs = qs.filter(year=int(year))
            except (ValueError, TypeError):
                pass

        # Check export size
        count = qs.count()
        if count > self.MAX_EXPORT_SIZE:
            return Response({
                "error": f"Export too large ({count} records). Maximum allowed: {self.MAX_EXPORT_SIZE}",
                "count": count,
                "max_allowed": self.MAX_EXPORT_SIZE,
                "suggestion": "Please apply more specific filters (department, week, year)."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if count == 0:
            return Response({
                "error": "No records found matching the specified filters."
            }, status=status.HTTP_404_NOT_FOUND)

        # Generate filename
        filename = f"performance_report_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        logger.info(
            f"Excel export initiated by {user.emp_id}: {count} records",
            extra={'user': user.emp_id, 'count': count}
        )
        
        return generate_excel_report(qs, filename)


# ===========================================================
# Individual PDF Report (Employee) with Authorization
# ===========================================================
class EmployeePerformancePDFView(generics.GenericAPIView):
    """
    Generate PDF performance report for a specific employee.
    
    Authorization:
        - Employees: Can view only their own report
        - Managers: Can view direct reports' reports
        - Admin: Can view any employee's report
    """
    permission_classes = [permissions.IsAuthenticated]

    @transaction.atomic
    def get(self, request, emp_id):
        """Generate PDF report with proper authorization."""
        user = request.user
        
        # Fetch employee with optimized query
        employee = get_object_or_404(
            Employee.objects.select_related(
                'user', 'department', 'manager__user'
            ),
            user__emp_id__iexact=emp_id
        )
        
        # Authorization checks
        if user.role == "Employee":
            # Employees can only view their own reports
            if employee.user != user:
                logger.warning(
                    f"Unauthorized PDF access attempt: {user.emp_id} tried to access {emp_id}"
                )
                return Response({
                    "error": "You can only view your own performance report."
                }, status=status.HTTP_403_FORBIDDEN)
                
        elif user.role == "Manager":
            # Managers can view their direct reports only
            if not (employee.manager and employee.manager.user == user):
                logger.warning(
                    f"Unauthorized PDF access: Manager {user.emp_id} tried to access {emp_id}"
                )
                return Response({
                    "error": "You can only view reports for your direct reports."
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Admin can view all (no check needed)
        
        # Load evaluations with optimization
        evaluations = PerformanceEvaluation.objects.select_related(
            'department', 'evaluator'
        ).filter(
            employee=employee
        ).order_by("-year", "-week_number")
        
        logger.info(
            f"PDF report generated for {emp_id} by {user.emp_id}",
            extra={'target': emp_id, 'requester': user.emp_id}
        )
        
        return generate_pdf_report(employee, evaluations)


# ===========================================================
# performance/utils_export.py (PRODUCTION-READY VERSION)
# ===========================================================
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from django.utils import timezone
import logging

logger = logging.getLogger("performance")


# ===========================================================
# Excel Export Utility (Memory-Optimized)
# ===========================================================
def generate_excel_report(evaluations, filename="performance_report.xlsx"):
    """
    Generate Excel report with memory-efficient streaming.
    Uses iterator() to process records in chunks for large datasets.
    
    Args:
        evaluations: QuerySet of PerformanceEvaluation objects
        filename: Output filename
    
    Returns:
        HttpResponse with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    # Define headers
    headers = [
        "Emp ID", "Employee Name", "Department", "Manager",
        "Week", "Year", "Total Score", "Average Score (%)", "Rank", "Remarks"
    ]
    ws.append(headers)

    # Format header row
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # Process evaluations in chunks (memory-efficient)
    row_count = 0
    error_count = 0
    
    for e in evaluations.iterator(chunk_size=1000):
        try:
            # Safe attribute access with fallbacks
            emp_name = f"{e.employee.user.first_name} {e.employee.user.last_name}".strip()
            if not emp_name:
                emp_name = e.employee.user.username
            
            dept_name = e.department.name if e.department else "-"
            
            # Safe manager access (checks FK ID first to avoid query)
            manager_name = "-"
            if e.employee.manager_id:
                try:
                    mgr = e.employee.manager.user
                    manager_name = f"{mgr.first_name} {mgr.last_name}".strip() or mgr.username
                except (AttributeError, TypeError):
                    manager_name = "-"
            
            # Append row data
            ws.append([
                e.employee.user.emp_id,
                emp_name,
                dept_name,
                manager_name,
                e.week_number,
                e.year,
                e.total_score,
                round(e.average_score, 2),
                e.rank,
                e.remarks or "",
            ])
            row_count += 1
            
        except Exception as ex:
            # Log error but continue processing
            logger.error(
                f"Error processing evaluation {e.id} for export: {ex}",
                exc_info=True
            )
            error_count += 1
            continue

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Set width with min/max bounds
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Generate HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    logger.info(
        f"Excel export generated: {row_count} records exported, {error_count} errors",
        extra={'rows': row_count, 'errors': error_count}
    )
    
    return response


# ===========================================================
# PDF Export Utility (Single Employee Report)
# ===========================================================
def generate_pdf_report(employee, evaluations):
    """
    Generate comprehensive PDF performance report for a single employee.
    Includes summary statistics and detailed evaluation history.
    
    Args:
        employee: Employee instance
        evaluations: QuerySet of PerformanceEvaluation objects
    
    Returns:
        HttpResponse with PDF file
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=f"{employee.user.emp_id}_Performance_Report",
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph("<b>Employee Performance Report</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    
    # Employee information with safe access
    try:
        emp_name = f"{employee.user.first_name} {employee.user.last_name}".strip()
        if not emp_name:
            emp_name = employee.user.username
        dept_name = employee.department.name if employee.department else "N/A"
    except (AttributeError, TypeError) as e:
        logger.error(f"Error accessing employee data for PDF: {e}")
        emp_name = "Unknown"
        dept_name = "N/A"
    
    # Employee details
    story.append(Paragraph(
        f"<b>Employee:</b> {emp_name} ({employee.user.emp_id})",
        styles["Normal"]
    ))
    story.append(Paragraph(f"<b>Department:</b> {dept_name}", styles["Normal"]))
    story.append(Paragraph(
        f"<b>Generated on:</b> {timezone.now().strftime('%d %b %Y, %H:%M')}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 12))

    # Convert evaluations to list
    eval_list = list(evaluations)
    
    if not eval_list:
        # No evaluations found
        story.append(Paragraph(
            "<i>No performance evaluations found for this employee.</i>",
            styles["Normal"]
        ))
    else:
        # Calculate summary statistics
        total_evals = len(eval_list)
        avg_score = sum(e.average_score for e in eval_list) / total_evals
        highest = max(e.average_score for e in eval_list)
        lowest = min(e.average_score for e in eval_list)
        
        # Summary section
        story.append(Paragraph("<b>Performance Summary</b>", styles["Heading2"]))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"<b>Total Evaluations:</b> {total_evals} | "
            f"<b>Average Score:</b> {round(avg_score, 2)}% | "
            f"<b>Highest:</b> {round(highest, 2)}% | "
            f"<b>Lowest:</b> {round(lowest, 2)}%",
            styles["Normal"]
        ))
        story.append(Spacer(1, 12))
        
        # Evaluation history table
        story.append(Paragraph("<b>Evaluation History</b>", styles["Heading2"]))
        story.append(Spacer(1, 6))
        
        # Table data
        data = [["Week", "Year", "Total Score", "Average (%)", "Rank", "Remarks"]]
        
        for e in eval_list:
            try:
                # Truncate long remarks
                remarks = str(e.remarks or "")
                if len(remarks) > 50:
                    remarks = remarks[:47] + "..."
                
                data.append([
                    str(e.week_number),
                    str(e.year),
                    str(e.total_score),
                    f"{round(e.average_score, 2)}%",
                    str(e.rank),
                    remarks,
                ])
            except Exception as ex:
                logger.error(f"Error processing evaluation {e.id} in PDF: {ex}")
                continue

        # Create and style table
        table = Table(data, repeatRows=1, colWidths=[50, 50, 80, 80, 50, 200])
        table.setStyle(TableStyle([
            # Header row styling
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            
            # Data rows styling
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            
            # Padding
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))

        story.append(table)

    # Build PDF
    try:
        doc.build(story)
    except Exception as e:
        logger.exception(f"Failed to build PDF for {employee.user.emp_id}: {e}")
        raise

    buffer.seek(0)

    # Generate HTTP response
    response = HttpResponse(buffer, content_type="application/pdf")
    filename = f"{employee.user.emp_id}_performance_report_{timezone.now().strftime('%Y%m%d')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    logger.info(
        f"PDF report generated for {employee.user.emp_id}: {len(eval_list)} evaluations",
        extra={'emp_id': employee.user.emp_id, 'eval_count': len(eval_list)}
    )
    
    return response


# ===========================================================
# Department Summary Statistics (Bonus Feature)
# ===========================================================
class DepartmentPerformanceSummaryView(generics.GenericAPIView):
    """
    Get aggregated performance statistics for a department.
    Returns: avg score, top performers, evaluation count, etc.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, dept_code):
        """Get department performance summary."""
        from django.db.models import Avg, Count, Max, Min
        
        user = request.user
        department = get_object_or_404(Department, code__iexact=dept_code)
        
        # Get time range from params
        year = request.query_params.get('year')
        week = request.query_params.get('week')
        
        qs = PerformanceEvaluation.objects.filter(department=department)
        
        if year:
            try:
                qs = qs.filter(year=int(year))
            except (ValueError, TypeError):
                pass
        
        if week:
            try:
                qs = qs.filter(week_number=int(week))
            except (ValueError, TypeError):
                pass
        
        # Calculate statistics
        stats = qs.aggregate(
            total_evaluations=Count('id'),
            avg_score=Avg('average_score'),
            highest_score=Max('average_score'),
            lowest_score=Min('average_score'),
        )
        
        # Get top 5 performers
        top_performers = qs.select_related(
            'employee__user'
        ).order_by('-average_score')[:5].values(
            'employee__user__emp_id',
            'employee__user__first_name',
            'employee__user__last_name',
            'average_score',
            'rank'
        )
        
        return Response({
            "department": {
                "code": department.code,
                "name": department.name
            },
            "statistics": {
                "total_evaluations": stats['total_evaluations'] or 0,
                "average_score": round(stats['avg_score'], 2) if stats['avg_score'] else 0,
                "highest_score": round(stats['highest_score'], 2) if stats['highest_score'] else 0,
                "lowest_score": round(stats['lowest_score'], 2) if stats['lowest_score'] else 0,
            },
            "top_performers": list(top_performers),
            "filters_applied": {
                "year": year,
                "week": week
            }
        })