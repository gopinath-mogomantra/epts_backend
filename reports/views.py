# ===============================================
# reports/views.py (Enhanced Version)
# ===============================================
# Handles:
# - Weekly Consolidated Report
# - Monthly Consolidated Report
# - Manager-Wise Report
# - Department-Wise Report
# - Employee Performance History
# - Excel Export (Weekly + Monthly)
# - PDF Export (Employee Performance Report)
# - Cached Report Management
# ===============================================

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from rest_framework.generics import ListAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.throttling import UserRateThrottle

from django.db.models import Avg, F, Q, Window, Prefetch, Count
from django.db.models.functions import Rank
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.core.cache import cache
from django.db import transaction
from django.views.decorators.cache import cache_page
from django.utils.decorators import method_decorator

from datetime import timedelta, datetime
from itertools import chain
from typing import Dict, List, Optional, Tuple, Any
import csv
import logging
import hashlib

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from employee.models import Employee
from performance.models import PerformanceEvaluation
from feedback.models import GeneralFeedback, ManagerFeedback, ClientFeedback
from .models import CachedReport
from .serializers import (
    WeeklyReportSerializer,
    MonthlyReportSerializer,
    EmployeeHistorySerializer,
    ManagerReportSerializer,
    DepartmentReportSerializer,
    CachedReportSerializer,
)

from reports.utils.pdf_generator import generate_employee_performance_pdf
from notifications.views import create_report_notification

logger = logging.getLogger(__name__)


# ===========================================================
# CUSTOM PERMISSIONS
# ===========================================================
class IsManagerOrAdmin(permissions.BasePermission):
    """Allow access only to managers and admins."""
    
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
            request.user.is_staff or 
            hasattr(request.user, 'employee') and request.user.employee.is_manager
        )


class CanViewDepartmentReports(permissions.BasePermission):
    """Allow users to view reports only for their department."""
    
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        
        # Admins can view all
        if request.user.is_staff:
            return True
        
        # Check if user has employee profile and department matches
        department_name = request.query_params.get("department_name")
        if not department_name:
            return True  # Let the view handle missing parameter
        
        try:
            employee = request.user.employee
            return employee.department and employee.department.name.lower() == department_name.lower()
        except AttributeError:
            return False


# ===========================================================
# CUSTOM THROTTLING
# ===========================================================
class ReportGenerationThrottle(UserRateThrottle):
    """Limit report generation to prevent system overload."""
    rate = '10/hour'


class ExportThrottle(UserRateThrottle):
    """Limit export operations."""
    rate = '20/hour'


# ===========================================================
# CUSTOM PAGINATION
# ===========================================================
class ReportPagination(PageNumberPagination):
    """Standard pagination for report listings."""
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 200


# ===========================================================
# HELPER FUNCTIONS (Enhanced with Caching & Optimization)
# ===========================================================
def get_cache_key(prefix: str, **kwargs) -> str:
    """Generate consistent cache keys."""
    sorted_items = sorted(kwargs.items())
    key_string = f"{prefix}:{':'.join(f'{k}={v}' for k, v in sorted_items)}"
    return hashlib.md5(key_string.encode()).hexdigest()


def get_feedback_average(
    employee: Employee, 
    start_date: Optional[datetime] = None, 
    end_date: Optional[datetime] = None,
    use_cache: bool = True
) -> float:
    """
    Compute average rating across all feedback sources for a given employee.
    
    Args:
        employee: Employee instance
        start_date: Optional start date filter
        end_date: Optional end date filter
        use_cache: Whether to use caching (default True)
        
    Returns:
        Average feedback rating rounded to 2 decimals
    """
    # Generate cache key
    if use_cache:
        cache_key = get_cache_key(
            'feedback_avg',
            emp_id=employee.id,
            start=start_date.isoformat() if start_date else 'all',
            end=end_date.isoformat() if end_date else 'all'
        )
        cached_result = cache.get(cache_key)
        if cached_result is not None:
            return cached_result
    
    filters = Q(employee=employee)
    if start_date and end_date:
        filters &= Q(created_at__range=(start_date, end_date))
    
    try:
        # Use list comprehension with prefetch for better performance
        ratings = list(chain(
            GeneralFeedback.objects.filter(filters).values_list("rating", flat=True),
            ManagerFeedback.objects.filter(filters).values_list("rating", flat=True),
            ClientFeedback.objects.filter(filters).values_list("rating", flat=True),
        ))
        
        result = round(sum(ratings) / len(ratings), 2) if ratings else 0.0
        
        # Cache for 15 minutes
        if use_cache:
            cache.set(cache_key, result, 900)
        
        return result
    except Exception as e:
        logger.error(f"Error calculating feedback average for employee {employee.id}: {e}")
        return 0.0


def get_feedback_batch(
    employee_ids: List[int], 
    start_date: Optional[datetime] = None, 
    end_date: Optional[datetime] = None
) -> Dict[int, float]:
    """
    Efficiently compute feedback averages for multiple employees.
    
    Args:
        employee_ids: List of employee IDs
        start_date: Optional start date filter
        end_date: Optional end date filter
        
    Returns:
        Dictionary mapping employee_id to average feedback rating
    """
    if not employee_ids:
        return {}
    
    filters = Q(employee_id__in=employee_ids)
    if start_date and end_date:
        filters &= Q(created_at__range=(start_date, end_date))
    
    # Aggregate all feedback types efficiently
    feedback_data: Dict[int, List[float]] = {}
    
    for model in [GeneralFeedback, ManagerFeedback, ClientFeedback]:
        for emp_id, rating in model.objects.filter(filters).values_list('employee_id', 'rating'):
            if emp_id not in feedback_data:
                feedback_data[emp_id] = []
            feedback_data[emp_id].append(float(rating))
    
    # Calculate averages
    return {
        emp_id: round(sum(ratings) / len(ratings), 2) if ratings else 0.0
        for emp_id, ratings in feedback_data.items()
    }


def validate_period_params(week: Optional[int] = None, month: Optional[int] = None, year: Optional[int] = None) -> Tuple[bool, str]:
    """
    Validate period parameters.
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    current_year = timezone.now().year
    
    if year is not None:
        if not (2000 <= year <= current_year + 1):
            return False, f"Year must be between 2000 and {current_year + 1}"
    
    if week is not None:
        if not (1 <= week <= 53):
            return False, "Week number must be between 1 and 53"
    
    if month is not None:
        if not (1 <= month <= 12):
            return False, "Month must be between 1 and 12"
    
    return True, ""


def send_report_notification_safe(user, report_type: str, link: str, message: str, department=None):
    """Safely send notification without breaking report generation."""
    try:
        create_report_notification(
            triggered_by=user,
            report_type=report_type,
            link=link,
            message=message,
            department=department,
        )
    except Exception as e:
        logger.error(f"Notification failed for {report_type}: {e}")


# ===========================================================
# BASE REPORT VIEW (DRY Pattern)
# ===========================================================
class BaseReportView(APIView):
    """Base class for report views with common functionality."""
    
    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [ReportGenerationThrottle]
    
    def get_week_year(self, request) -> Tuple[int, int]:
        """Extract and validate week and year from request."""
        now = timezone.now()
        week = int(request.query_params.get("week", now.isocalendar()[1]))
        year = int(request.query_params.get("year", now.year))
        
        is_valid, error_msg = validate_period_params(week=week, year=year)
        if not is_valid:
            raise ValueError(error_msg)
        
        return week, year
    
    def get_month_year(self, request) -> Tuple[int, int]:
        """Extract and validate month and year from request."""
        now = timezone.now()
        month = int(request.query_params.get("month", now.month))
        year = int(request.query_params.get("year", now.year))
        
        is_valid, error_msg = validate_period_params(month=month, year=year)
        if not is_valid:
            raise ValueError(error_msg)
        
        return month, year
    
    def should_cache(self, request) -> bool:
        """Check if report should be cached."""
        return request.query_params.get("save_cache", "false").lower() == "true"
    
    def handle_error(self, error: Exception, context: str) -> Response:
        """Standardized error handling."""
        logger.exception(f"{context}: {str(error)}")
        return Response(
            {"error": str(error)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ===========================================================
# 1. WEEKLY CONSOLIDATED REPORT (Enhanced)
# ===========================================================
class WeeklyReportView(BaseReportView):
    """
    Generate consolidated weekly performance report.
    
    Query Parameters:
        - week (int): Week number (1-53), default: current week
        - year (int): Year, default: current year
        - save_cache (bool): Whether to cache results, default: false
        
    Returns:
        JSON with weekly performance records including rankings
    """
    
    def get(self, request):
        try:
            week, year = self.get_week_year(request)
            save_cache = self.should_cache(request)
            
            # Check cache first
            cache_key = get_cache_key('weekly_report', week=week, year=year)
            cached_data = cache.get(cache_key)
            if cached_data and not save_cache:
                logger.info(f"Returning cached weekly report for Week {week}, {year}")
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Optimized query with select_related and prefetch
            qs = (
                PerformanceEvaluation.objects
                .filter(week_number=week, year=year)
                .select_related("employee__user", "employee__department", "department")
                .annotate(emp_id=F("employee__user__emp_id"))
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )
            
            # Batch fetch feedback averages
            employee_ids = list(qs.values_list("employee_id", flat=True))
            feedback_map = get_feedback_batch(employee_ids)
            
            # Annotate with ranking
            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(), 
                    order_by=F("total_score").desc()
                )
            )
            
            # Build result set
            result = [
                {
                    "emp_id": p.employee.user.emp_id,
                    "employee_full_name": f"{p.employee.user.first_name} {p.employee.user.last_name}".strip() or "N/A",
                    "department": p.department.name if p.department else "-",
                    "total_score": float(p.total_score),
                    "average_score": float(p.average_score),
                    "feedback_avg": feedback_map.get(p.employee.id, 0.0),
                    "week_number": week,
                    "year": year,
                    "rank": int(p.computed_rank),
                    "remarks": p.remarks or "",
                }
                for p in ranked
            ]
            
            # Serialize data
            serialized_data = WeeklyReportSerializer(result, many=True).data
            
            response_data = {
                "evaluation_period": f"Week {week}, {year}",
                "total_records": len(result),
                "records": serialized_data,
            }
            
            # Cache the response (15 minutes)
            cache.set(cache_key, response_data, 900)
            
            # Save to database if requested
            if save_cache:
                with transaction.atomic():
                    CachedReport.objects.update_or_create(
                        report_type="weekly",
                        year=year,
                        week_number=week,
                        defaults={
                            "payload": {"records": result},
                            "generated_by": request.user,
                        },
                    )
            
            # Send notification
            send_report_notification_safe(
                user=request.user,
                report_type="Weekly Report",
                link=f"/reports/weekly/?week={week}&year={year}",
                message=f"Weekly performance report generated for Week {week}, {year}.",
                department=None,
            )
            
            logger.info(f"Weekly report generated for Week {week}, {year} by {request.user.username}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "WeeklyReport Error")


# ===========================================================
# 2. MONTHLY CONSOLIDATED REPORT (Enhanced)
# ===========================================================
class MonthlyReportView(BaseReportView):
    """
    Generate monthly aggregated performance report.
    
    Query Parameters:
        - month (int): Month number (1-12), default: current month
        - year (int): Year, default: current year
        - save_cache (bool): Whether to cache results, default: false
        
    Returns:
        JSON with monthly performance summaries including best weeks
    """
    
    def get(self, request):
        try:
            month, year = self.get_month_year(request)
            save_cache = self.should_cache(request)
            
            # Check cache first
            cache_key = get_cache_key('monthly_report', month=month, year=year)
            cached_data = cache.get(cache_key)
            if cached_data and not save_cache:
                logger.info(f"Returning cached monthly report for {month}/{year}")
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Optimized query
            qs = (
                PerformanceEvaluation.objects
                .filter(review_date__month=month, year=year)
                .select_related("employee__user", "employee__department", "department")
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for {month}/{year}."},
                    status=status.HTTP_200_OK,
                )
            
            data = []
            employee_ids = list(qs.values_list("employee_id", flat=True).distinct())
            employees = Employee.objects.filter(id__in=employee_ids).select_related("user", "department")
            
            for emp in employees:
                emp_qs = qs.filter(employee=emp)
                if not emp_qs.exists():
                    continue
                
                # Aggregate scores
                avg_data = emp_qs.aggregate(avg=Avg("average_score"))
                avg_score = round(avg_data["avg"], 2) if avg_data["avg"] else 0.0
                
                # Find best week
                best_week_obj = emp_qs.order_by("-average_score").first()
                
                if not best_week_obj:
                    continue
                
                # Calculate feedback average for the month
                start_date = datetime(year, month, 1)
                if month == 12:
                    end_date = datetime(year + 1, 1, 1)
                else:
                    end_date = datetime(year, month + 1, 1)
                
                fb_avg = get_feedback_average(emp, start_date=start_date, end_date=end_date)
                
                data.append({
                    "emp_id": emp.user.emp_id,
                    "employee_full_name": f"{emp.user.first_name} {emp.user.last_name}".strip() or "N/A",
                    "department": emp.department.name if emp.department else "-",
                    "month": month,
                    "year": year,
                    "avg_score": avg_score,
                    "feedback_avg": fb_avg,
                    "best_week": best_week_obj.week_number,
                    "best_week_score": float(best_week_obj.average_score),
                })
            
            # Serialize data
            serialized_data = MonthlyReportSerializer(data, many=True).data
            
            response_data = {
                "evaluation_period": f"Month {month}, {year}",
                "total_records": len(data),
                "records": serialized_data,
            }
            
            # Cache the response (30 minutes)
            cache.set(cache_key, response_data, 1800)
            
            # Save to database if requested
            if save_cache:
                with transaction.atomic():
                    CachedReport.objects.update_or_create(
                        report_type="monthly",
                        year=year,
                        month=month,
                        defaults={
                            "payload": {"records": data},
                            "generated_by": request.user,
                        },
                    )
            
            # Send notification
            send_report_notification_safe(
                user=request.user,
                report_type="Monthly Report",
                link=f"/reports/monthly/?month={month}&year={year}",
                message=f"Monthly performance report generated for {month}/{year}.",
                department=None,
            )
            
            logger.info(f"Monthly report generated for {month}/{year} by {request.user.username}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "MonthlyReport Error")


# ===========================================================
# 3. DEPARTMENT-WISE WEEKLY REPORT (Enhanced)
# ===========================================================
class DepartmentReportView(BaseReportView):
    """
    Generate department-wise weekly performance report.
    
    Query Parameters:
        - department_name (str): Department name (required)
        - week (int): Week number (1-53), default: current week
        - year (int): Year, default: current year
        
    Returns:
        JSON with department performance records
    """
    
    permission_classes = [permissions.IsAuthenticated, CanViewDepartmentReports]
    
    def get(self, request):
        try:
            department_name = request.query_params.get("department_name")
            if not department_name:
                return Response(
                    {"error": "Please provide department_name parameter."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            week, year = self.get_week_year(request)
            
            # Check cache
            cache_key = get_cache_key('dept_report', dept=department_name, week=week, year=year)
            cached_data = cache.get(cache_key)
            if cached_data:
                logger.info(f"Returning cached department report for {department_name}")
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Optimized query
            employees = Employee.objects.filter(
                department__name__iexact=department_name
            ).select_related("user", "department")
            
            if not employees.exists():
                return Response(
                    {"message": f"No employees found in department '{department_name}'."},
                    status=status.HTTP_200_OK
                )
            
            qs = (
                PerformanceEvaluation.objects
                .filter(employee__in=employees, week_number=week, year=year)
                .select_related("employee__user", "employee__department", "department")
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for department '{department_name}' in Week {week}, {year}."},
                    status=status.HTTP_200_OK
                )
            
            # Batch fetch feedback
            employee_ids = list(qs.values_list("employee_id", flat=True))
            feedback_map = get_feedback_batch(employee_ids)
            
            # Annotate with ranking
            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc()
                )
            )
            
            # Build records
            records = [
                {
                    "department_name": department_name,
                    "emp_id": perf.employee.user.emp_id,
                    "employee_full_name": f"{perf.employee.user.first_name} {perf.employee.user.last_name}".strip() or "N/A",
                    "manager_full_name": f"{perf.employee.manager.user.first_name} {perf.employee.manager.user.last_name}".strip() if perf.employee.manager else "N/A",
                    "total_score": float(perf.total_score),
                    "average_score": float(perf.average_score),
                    "feedback_avg": feedback_map.get(perf.employee.id, 0.0),
                    "week_number": week,
                    "year": year,
                    "rank": int(perf.computed_rank),
                    "remarks": perf.remarks or "",
                }
                for perf in ranked
            ]
            
            response_data = {
                "department_name": department_name,
                "evaluation_period": f"Week {week}, {year}",
                "total_employees": len(records),
                "records": records,
            }
            
            # Cache for 15 minutes
            cache.set(cache_key, response_data, 900)
            
            # Send notification
            send_report_notification_safe(
                user=request.user,
                report_type="Department Weekly Report",
                link=f"/reports/department/?department_name={department_name}&week={week}&year={year}",
                message=f"Department-wise report generated for {department_name} (Week {week}, {year}).",
                department=None,
            )
            
            logger.info(f"Department report generated for {department_name}, Week {week}, {year}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "DepartmentReport Error")


# ===========================================================
# 4. MANAGER REPORT (Enhanced Implementation)
# ===========================================================
class ManagerReportView(BaseReportView):
    """
    Generate manager-wise weekly performance report.
    
    Query Parameters:
        - manager_id (str): Manager's employee ID (optional, defaults to current user)
        - week (int): Week number (1-53), default: current week
        - year (int): Year, default: current year
        
    Returns:
        JSON with performance records of all team members under the manager
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    
    def get(self, request):
        try:
            manager_id = request.query_params.get("manager_id")
            week, year = self.get_week_year(request)
            
            # Determine manager
            if manager_id:
                try:
                    manager = Employee.objects.select_related("user").get(user__emp_id=manager_id)
                except Employee.DoesNotExist:
                    return Response(
                        {"error": f"Manager with ID {manager_id} not found."},
                        status=status.HTTP_404_NOT_FOUND
                    )
            else:
                # Use current user as manager
                try:
                    manager = request.user.employee
                except AttributeError:
                    return Response(
                        {"error": "Current user does not have an employee profile."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Check cache
            cache_key = get_cache_key('manager_report', mgr_id=manager.id, week=week, year=year)
            cached_data = cache.get(cache_key)
            if cached_data:
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Get team members
            team_members = Employee.objects.filter(manager=manager).select_related("user", "department")
            
            if not team_members.exists():
                return Response(
                    {"message": f"No team members found under manager {manager.user.emp_id}."},
                    status=status.HTTP_200_OK
                )
            
            # Get performance records
            qs = (
                PerformanceEvaluation.objects
                .filter(employee__in=team_members, week_number=week, year=year)
                .select_related("employee__user", "employee__department", "department")
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for team in Week {week}, {year}."},
                    status=status.HTTP_200_OK
                )
            
            # Batch feedback
            employee_ids = list(qs.values_list("employee_id", flat=True))
            feedback_map = get_feedback_batch(employee_ids)
            
            # Rank
            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc()
                )
            )
            
            # Build records
            records = [
                {
                    "manager_full_name": f"{manager.user.first_name} {manager.user.last_name}".strip(),
                    "emp_id": perf.employee.user.emp_id,
                    "employee_full_name": f"{perf.employee.user.first_name} {perf.employee.user.last_name}".strip(),
                    "department": perf.department.name if perf.department else "-",
                    "total_score": float(perf.total_score),
                    "average_score": float(perf.average_score),
                    "feedback_avg": feedback_map.get(perf.employee.id, 0.0),
                    "week_number": week,
                    "year": year,
                    "rank": int(perf.computed_rank),
                    "remarks": perf.remarks or "",
                }
                for perf in ranked
            ]
            
            response_data = {
                "manager_name": f"{manager.user.first_name} {manager.user.last_name}".strip(),
                "manager_id": manager.user.emp_id,
                "evaluation_period": f"Week {week}, {year}",
                "total_team_members": len(records),
                "records": ManagerReportSerializer(records, many=True).data,
            }
            
            # Cache for 15 minutes
            cache.set(cache_key, response_data, 900)
            
            # Notification
            send_report_notification_safe(
                user=request.user,
                report_type="Manager Report",
                link=f"/reports/manager/?manager_id={manager.user.emp_id}&week={week}&year={year}",
                message=f"Manager report generated for {manager.user.first_name} {manager.user.last_name} (Week {week}, {year}).",
                department=manager.department,
            )
            
            logger.info(f"Manager report generated for {manager.user.emp_id}, Week {week}, {year}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "ManagerReport Error")


# ===========================================================
# 5. EXCEL EXPORT VIEWS (Enhanced with Error Handling)
# ===========================================================
class BaseExcelExportView(APIView):
    """Base class for Excel export functionality."""
    
    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [ExportThrottle]
    
    def apply_excel_styling(self, ws, header_color="4472C4"):
        """Apply consistent styling to Excel worksheet."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Apply to header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 characters
            ws.column_dimensions[col_letter].width = adjusted_width


class ExportWeeklyExcelView(BaseExcelExportView):
    """
    Export weekly performance data to Excel format.
    
    Query Parameters:
        - week (int): Week number (1-53), default: current week
        - year (int): Year, default: current year
        
    Returns:
        Excel file download with formatted weekly report
    """
    
    def get(self, request):
        try:
            now = timezone.now()
            week = int(request.query_params.get("week", now.isocalendar()[1]))
            year = int(request.query_params.get("year", now.year))
            
            # Validate parameters
            is_valid, error_msg = validate_period_params(week=week, year=year)
            if not is_valid:
                return Response({"error": error_msg}, status=status.HTTP_400_BAD_REQUEST)
            
            # Fetch data with optimized query
            qs = (
                PerformanceEvaluation.objects
                .filter(week_number=week, year=year)
                .select_related("employee__user", "employee__department", "department")
                .annotate(emp_id=F("employee__user__emp_id"))
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Week_{week}_{year}"
            
            # Add headers
            headers = [
                "Emp ID", "Employee Name", "Department",
                "Total Score", "Average Score", "Feedback Avg",
                "Rank", "Remarks"
            ]
            ws.append(headers)
            
            # Apply styling
            self.apply_excel_styling(ws, header_color="4472C4")
            
            # Fetch feedback data in batch
            employee_ids = list(qs.values_list("employee_id", flat=True))
            feedback_map = get_feedback_batch(employee_ids)
            
            # Add ranked data
            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc()
                )
            )
            
            for perf in ranked:
                ws.append([
                    perf.employee.user.emp_id,
                    f"{perf.employee.user.first_name} {perf.employee.user.last_name}".strip(),
                    perf.department.name if perf.department else "-",
                    float(perf.total_score),
                    float(perf.average_score),
                    feedback_map.get(perf.employee.id, 0.0),
                    int(perf.computed_rank),
                    perf.remarks or "",
                ])
            
            # Auto-adjust columns
            self.auto_adjust_columns(ws)
            
            # Create response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"Weekly_Performance_Report_Week{week}_{year}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            
            logger.info(f"Weekly Excel export generated for Week {week}, {year} by {request.user.username}")
            return response
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception("ExportWeeklyExcel Error: %s", str(e))
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportMonthlyExcelView(BaseExcelExportView):
    """
    Export monthly performance summary to Excel format.
    
    Query Parameters:
        - month (int): Month number (1-12), default: current month
        - year (int): Year, default: current year
        
    Returns:
        Excel file download with formatted monthly report
    """
    
    def get(self, request):
        try:
            now = timezone.now()
            month = int(request.query_params.get("month", now.month))
            year = int(request.query_params.get("year", now.year))
            
            # Validate parameters
            is_valid, error_msg = validate_period_params(month=month, year=year)
            if not is_valid:
                return Response({"error": error_msg}, status=status.HTTP_400_BAD_REQUEST)
            
            # Fetch data
            qs = (
                PerformanceEvaluation.objects
                .filter(review_date__month=month, year=year)
                .select_related("employee__user", "employee__department", "department")
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for {month}/{year}."},
                    status=status.HTTP_200_OK,
                )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Month_{month}_{year}"
            
            # Add headers
            headers = [
                "Emp ID", "Employee Name", "Department",
                "Average Score", "Feedback Avg",
                "Best Week", "Best Week Score"
            ]
            ws.append(headers)
            
            # Apply styling with green header
            self.apply_excel_styling(ws, header_color="70AD47")
            
            # Calculate monthly aggregates
            employee_ids = list(qs.values_list("employee_id", flat=True).distinct())
            employees = Employee.objects.filter(id__in=employee_ids).select_related("user", "department")
            
            # Date range for feedback
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
            
            feedback_map = get_feedback_batch(employee_ids, start_date=start_date, end_date=end_date)
            
            for emp in employees:
                emp_qs = qs.filter(employee=emp)
                if not emp_qs.exists():
                    continue
                
                avg_data = emp_qs.aggregate(avg=Avg("average_score"))
                avg_score = round(avg_data["avg"], 2) if avg_data["avg"] else 0.0
                
                best_week_obj = emp_qs.order_by("-average_score").first()
                if not best_week_obj:
                    continue
                
                ws.append([
                    emp.user.emp_id,
                    f"{emp.user.first_name} {emp.user.last_name}".strip(),
                    emp.department.name if emp.department else "-",
                    avg_score,
                    feedback_map.get(emp.id, 0.0),
                    best_week_obj.week_number,
                    float(best_week_obj.average_score),
                ])
            
            # Auto-adjust columns
            self.auto_adjust_columns(ws)
            
            # Create response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"Monthly_Performance_Report_{month}_{year}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            
            logger.info(f"Monthly Excel export generated for {month}/{year} by {request.user.username}")
            return response
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception("ExportMonthlyExcel Error: %s", str(e))
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# 6. CSV EXPORT VIEW (New Addition)
# ===========================================================
class ExportWeeklyCSVView(BaseReportView):
    """
    Export weekly performance data to CSV format.
    
    Query Parameters:
        - week (int): Week number (1-53), default: current week
        - year (int): Year, default: current year
        
    Returns:
        CSV file download
    """
    
    throttle_classes = [ExportThrottle]
    
    def get(self, request):
        try:
            week, year = self.get_week_year(request)
            
            # Fetch data
            qs = (
                PerformanceEvaluation.objects
                .filter(week_number=week, year=year)
                .select_related("employee__user", "employee__department", "department")
            )
            
            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            filename = f"Weekly_Performance_Report_Week{week}_{year}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            
            # Write headers
            writer.writerow([
                'Emp ID', 'Employee Name', 'Department',
                'Total Score', 'Average Score', 'Feedback Avg',
                'Rank', 'Remarks'
            ])
            
            # Fetch feedback
            employee_ids = list(qs.values_list("employee_id", flat=True))
            feedback_map = get_feedback_batch(employee_ids)
            
            # Write data
            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc()
                )
            )
            
            for perf in ranked:
                writer.writerow([
                    perf.employee.user.emp_id,
                    f"{perf.employee.user.first_name} {perf.employee.user.last_name}".strip(),
                    perf.department.name if perf.department else "-",
                    float(perf.total_score),
                    float(perf.average_score),
                    feedback_map.get(perf.employee.id, 0.0),
                    int(perf.computed_rank),
                    perf.remarks or "",
                ])
            
            logger.info(f"CSV export generated for Week {week}, {year}")
            return response
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "ExportWeeklyCSV Error")


# ===========================================================
# 7. PDF EXPORT (Enhanced)
# ===========================================================
class PrintPerformanceReportView(BaseReportView):
    """
    Generate and return a downloadable PDF report for individual employee's weekly performance.
    
    Path Parameters:
        - emp_id (str): Employee ID
        
    Query Parameters:
        - week (int): Week number (1-53), default: current week
        - year (int): Year, default: current year
        
    Returns:
        PDF file download
        
    Example:
        GET /api/reports/print/EMP0001/?week=44&year=2025
    """
    
    throttle_classes = [ExportThrottle]
    
    def get(self, request, emp_id):
        try:
            week, year = self.get_week_year(request)
            
            # Validate and fetch employee
            try:
                employee = Employee.objects.select_related(
                    "user", "department", "manager__user"
                ).get(user__emp_id__iexact=emp_id)
            except Employee.DoesNotExist:
                return Response(
                    {"error": f"Employee with ID {emp_id} not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Permission check: Users can only view their own reports unless they're managers/admins
            if not (request.user.is_staff or 
                    request.user.emp_id == emp_id or 
                    (hasattr(request.user, 'employee') and 
                     request.user.employee.is_manager)):
                return Response(
                    {"error": "You don't have permission to view this report."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Fetch performance evaluations
            evaluations = (
                PerformanceEvaluation.objects
                .filter(employee=employee, week_number=week, year=year)
                .select_related("employee__user", "department", "reviewer__user")
                .order_by("-created_at")
            )
            
            if not evaluations.exists():
                return Response(
                    {"message": f"No performance records found for {emp_id} in Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )
            
            # Compute feedback average
            employee.latest_feedback_avg = get_feedback_average(employee, use_cache=True)
            
            # Generate PDF
            pdf_response = generate_employee_performance_pdf(
                employee=employee,
                evaluations=evaluations,
                week=f"Week {week}, {year}"
            )
            
            # Send notification
            send_report_notification_safe(
                user=request.user,
                report_type="Employee Performance PDF",
                link=request.get_full_path(),
                message=f"PDF performance report generated for {employee.user.emp_id} ({employee.user.first_name} {employee.user.last_name}).",
                department=employee.department,
            )
            
            logger.info(f"PDF report generated for {emp_id}, Week {week}, {year} by {request.user.username}")
            return pdf_response
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "PrintPerformanceReport Error")


# ===========================================================
# 8. EMPLOYEE PERFORMANCE HISTORY (New Addition)
# ===========================================================
class EmployeeHistoryView(BaseReportView):
    """
    Retrieve performance history for a specific employee across multiple weeks.
    
    Path Parameters:
        - emp_id (str): Employee ID
        
    Query Parameters:
        - start_week (int): Starting week number
        - end_week (int): Ending week number
        - year (int): Year, default: current year
        
    Returns:
        JSON with historical performance data
    """
    
    def get(self, request, emp_id):
        try:
            year = int(request.query_params.get("year", timezone.now().year))
            start_week = int(request.query_params.get("start_week", 1))
            end_week = int(request.query_params.get("end_week", timezone.now().isocalendar()[1]))
            
            # Validate parameters
            if not (1 <= start_week <= 53 and 1 <= end_week <= 53):
                return Response(
                    {"error": "Week numbers must be between 1 and 53."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if start_week > end_week:
                return Response(
                    {"error": "start_week must be less than or equal to end_week."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Fetch employee
            try:
                employee = Employee.objects.select_related("user", "department").get(
                    user__emp_id__iexact=emp_id
                )
            except Employee.DoesNotExist:
                return Response(
                    {"error": f"Employee with ID {emp_id} not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Check cache
            cache_key = get_cache_key(
                'emp_history',
                emp_id=emp_id,
                start=start_week,
                end=end_week,
                year=year
            )
            cached_data = cache.get(cache_key)
            if cached_data:
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Fetch performance data
            evaluations = (
                PerformanceEvaluation.objects
                .filter(
                    employee=employee,
                    week_number__gte=start_week,
                    week_number__lte=end_week,
                    year=year
                )
                .select_related("department")
                .order_by("week_number")
            )
            
            if not evaluations.exists():
                return Response(
                    {"message": f"No performance history found for {emp_id} between weeks {start_week}-{end_week}, {year}."},
                    status=status.HTTP_200_OK
                )
            
            # Build history records
            history = []
            for eval in evaluations:
                fb_avg = get_feedback_average(employee, use_cache=True)
                history.append({
                    "week_number": eval.week_number,
                    "year": year,
                    "average_score": float(eval.average_score),
                    "feedback_avg": fb_avg,
                    "remarks": eval.remarks or "",
                    "rank": None,  # Could calculate if needed
                })
            
            response_data = {
                "employee_id": emp_id,
                "employee_name": f"{employee.user.first_name} {employee.user.last_name}".strip(),
                "department": employee.department.name if employee.department else "-",
                "period": f"Weeks {start_week}-{end_week}, {year}",
                "total_weeks": len(history),
                "history": EmployeeHistorySerializer(history, many=True).data,
            }
            
            # Cache for 1 hour
            cache.set(cache_key, response_data, 3600)
            
            logger.info(f"Performance history retrieved for {emp_id}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return self.handle_error(e, "EmployeeHistory Error")


# ===========================================================
# 9. CACHED REPORT MANAGEMENT (Enhanced)
# ===========================================================
class CachedReportListView(ListAPIView):
    """
    List all cached reports with filtering and pagination.
    
    Query Parameters:
        - report_type (str): Filter by report type
        - year (int): Filter by year
        - is_archived (bool): Include archived reports
        
    Returns:
        Paginated list of cached reports
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    serializer_class = CachedReportSerializer
    pagination_class = ReportPagination
    
    def get_queryset(self):
        queryset = CachedReport.objects.select_related(
            "generated_by", "manager", "department"
        ).order_by("-created_at")
        
        # Apply filters
        report_type = self.request.query_params.get("report_type")
        if report_type:
            queryset = queryset.filter(report_type__iexact=report_type)
        
        year = self.request.query_params.get("year")
        if year:
            try:
                queryset = queryset.filter(year=int(year))
            except ValueError:
                pass
        
        is_archived = self.request.query_params.get("is_archived", "false").lower()
        if is_archived != "true":
            queryset = queryset.filter(is_archived=False)
        
        return queryset


class CachedReportDetailView(APIView):
    """
    Retrieve a specific cached report by ID.
    
    Path Parameters:
        - pk (int): Cached report ID
        
    Returns:
        Detailed cached report data
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    
    def get(self, request, pk):
        try:
            report = get_object_or_404(
                CachedReport.objects.select_related("generated_by", "manager", "department"),
                pk=pk
            )
            serializer = CachedReportSerializer(report)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"CachedReportDetail Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CachedReportArchiveView(APIView):
    """
    Archive a cached report (soft delete).
    
    Path Parameters:
        - pk (int): Cached report ID
        
    Returns:
        Success message
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    
    def post(self, request, pk):
        try:
            report = get_object_or_404(CachedReport, pk=pk)
            
            with transaction.atomic():
                report.is_archived = True
                report.save(update_fields=["is_archived", "updated_at"])
            
            logger.info(f"Cached report {report.id} archived by {request.user.username}")
            return Response(
                {"message": f"Report {report.id} archived successfully."},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            logger.exception(f"CachedReportArchive Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CachedReportRestoreView(APIView):
    """
    Restore an archived cached report.
    
    Path Parameters:
        - pk (int): Cached report ID
        
    Returns:
        Success message
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    
    def post(self, request, pk):
        try:
            report = get_object_or_404(CachedReport, pk=pk)
            
            with transaction.atomic():
                report.is_archived = False
                report.save(update_fields=["is_archived", "updated_at"])
            
            logger.info(f"Cached report {report.id} restored by {request.user.username}")
            return Response(
                {"message": f"Report {report.id} restored successfully."},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            logger.exception(f"CachedReportRestore Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CachedReportDeleteView(APIView):
    """
    Permanently delete a cached report (hard delete).
    
    Path Parameters:
        - pk (int): Cached report ID
        
    Returns:
        Success message
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    
    def delete(self, request, pk):
        try:
            report = get_object_or_404(CachedReport, pk=pk)
            report_id = report.id
            
            with transaction.atomic():
                report.delete()
            
            logger.warning(f"Cached report {report_id} permanently deleted by {request.user.username}")
            return Response(
                {"message": f"Report {report_id} permanently deleted."},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            logger.exception(f"CachedReportDelete Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# 10. REPORT STATISTICS VIEW (New Addition)
# ===========================================================
class ReportStatisticsView(APIView):
    """
    Get statistical overview of reports.
    
    Returns:
        Summary statistics including:
        - Total reports generated
        - Reports by type
        - Recent activity
        - Most active users
    """
    
    permission_classes = [permissions.IsAuthenticated, IsManagerOrAdmin]
    
    @method_decorator(cache_page(300))  # Cache for 5 minutes
    def get(self, request):
        try:
            # Get counts by report type
            report_counts = (
                CachedReport.objects
                .filter(is_archived=False)
                .values('report_type')
                .annotate(count=Count('id'))
            )
            
            # Recent reports (last 7 days)
            seven_days_ago = timezone.now() - timedelta(days=7)
            recent_count = CachedReport.objects.filter(
                created_at__gte=seven_days_ago,
                is_archived=False
            ).count()
            
            # Most active users
            top_users = (
                CachedReport.objects
                .filter(is_archived=False)
                .values('generated_by__username', 'generated_by__first_name', 'generated_by__last_name')
                .annotate(report_count=Count('id'))
                .order_by('-report_count')[:5]
            )
            
            statistics = {
                "total_reports": CachedReport.objects.filter(is_archived=False).count(),
                "total_archived": CachedReport.objects.filter(is_archived=True).count(),
                "reports_by_type": {item['report_type']: item['count'] for item in report_counts},
                "recent_reports_7days": recent_count,
                "top_generators": [
                    {
                        "username": user['generated_by__username'],
                        "name": f"{user['generated_by__first_name']} {user['generated_by__last_name']}".strip(),
                        "report_count": user['report_count']
                    }
                    for user in top_users
                ],
            }
            
            return Response(statistics, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception(f"ReportStatistics Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)